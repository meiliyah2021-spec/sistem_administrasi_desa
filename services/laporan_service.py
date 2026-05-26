from database.oracle_helper import execute_query
from datetime import datetime

class LaporanService:
    """Provides reporting engine filtering the view laporan_pengajuan_layanan."""

    @staticmethod
    def get_laporan(tgl_awal=None, tgl_akhir=None, id_status=None, id_layanan=None, search_name_nik=None):
        """
        Queries the reporting database view 'laporan_pengajuan_layanan' with dynamic filters.
        """
        sql = "SELECT * FROM laporan_pengajuan_layanan WHERE 1=1"
        params = {}

        if tgl_awal:
            # Parse & bind date
            sql += " AND tanggal_pengajuan >= TO_DATE(:tgl_awal, 'YYYY-MM-DD')"
            params["tgl_awal"] = tgl_awal
            
        if tgl_akhir:
            sql += " AND tanggal_pengajuan <= TO_DATE(:tgl_akhir, 'YYYY-MM-DD') + 1"
            params["tgl_akhir"] = tgl_akhir
            
        if id_status:
            sql += " AND id_status = :id_status"
            params["id_status"] = int(id_status)
            
        if id_layanan:
            sql += " AND id_layanan = :id_layanan"
            params["id_layanan"] = int(id_layanan)
            
        if search_name_nik:
            sql += " AND (UPPER(nama_penduduk) LIKE :search OR UPPER(nik) LIKE :search)"
            params["search"] = f"%{search_name_nik.upper()}%"

        sql += " ORDER BY tanggal_pengajuan DESC"
        return execute_query(sql, params)

    @staticmethod
    def export_excel(tgl_awal=None, tgl_akhir=None, id_status=None, id_layanan=None, search_name_nik=None):
        """Generates an in-memory Workbook representing the report data."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        # Query data
        records = LaporanService.get_laporan(tgl_awal, tgl_akhir, id_status, id_layanan, search_name_nik)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Pengajuan Desa"

        # Style Definitions
        title_font = Font(name="Arial", size=16, bold=True, color="4B0082")
        subheader_font = Font(name="Arial", size=10, italic=True, color="666666")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        header_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")
        
        thin_border_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 1. Title Header Block
        ws.merge_cells("A1:I1")
        ws["A1"] = "LAPORAN PENGAJUAN LAYANAN ADMINISTRASI KEPENDUDUKAN DESA"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:I2")
        tgl_cetak = datetime.now().strftime("%d-%m-%Y %H:%M")
        ws["A2"] = f"Dicetak pada: {tgl_cetak} | Filter: Tgl Awal ({tgl_awal or 'Semua'}), Tgl Akhir ({tgl_akhir or 'Semua'})"
        ws["A2"].font = subheader_font
        ws["A2"].alignment = center_align
        ws.row_dimensions[2].height = 20

        # Blank spacing row
        ws.row_dimensions[3].height = 10

        # Header Columns
        headers = [
            "ID Pengajuan", "Tanggal", "NIK Pemohon", "Nama Pemohon", 
            "No. KK", "Alamat Warga", "Layanan Kependudukan", "Status", "Petugas Verifikator"
        ]
        
        cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        ws.row_dimensions[4].height = 25
        
        for col, h in zip(cols, headers):
            cell = ws[f"{col}4"]
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Populate rows
        row_idx = 5
        for idx, rec in enumerate(records):
            ws.row_dimensions[row_idx].height = 22
            
            row_data = [
                rec['id_pengajuan'],
                rec['tanggal_pengajuan'].strftime("%Y-%m-%d %H:%M") if isinstance(rec['tanggal_pengajuan'], datetime) else str(rec['tanggal_pengajuan']),
                rec['nik'],
                rec['nama_penduduk'],
                rec['no_kk'],
                rec['alamat'],
                rec['layanan'],
                rec['status'],
                rec['petugas']
            ]
            
            for col, val in zip(cols, row_data):
                cell = ws[f"{col}{row_idx}"]
                cell.value = val
                cell.font = data_font
                cell.border = thin_border
                
                if col in ["A", "B", "C", "E", "H"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            row_idx += 1

        # Adjust dimensions based on contents
        for col in cols:
            max_len = 0
            for row in range(4, row_idx):
                val = str(ws[f"{col}{row}"].value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col].width = max(max_len + 4, 12)

        # Save to virtual stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
